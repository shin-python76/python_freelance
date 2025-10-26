# 家計簿：カテゴリ別合計（最小コード）
# 前提：samplecode/sample_created.xlsx の「テストシート」に
#   A1:D1 = 日付 / カテゴリ / 内容 / 金額 のヘッダ、
#   2行目以降にデータがある想定。
# 出力：同じブック内に「カテゴリ別合計」シートを新規作成して書き出す。

from pathlib import Path
from openpyxl import load_workbook

# ---- 1) 入力ブックとシート名を指定（必要ならここだけ直す） ----
FILE = Path("samplecode/sample_created.xlsx")
INPUT_SHEET = "テストシート"
OUTPUT_SHEET = "カテゴリ別合計"

# ---- 2) ブックを開いて、入力シートを取得 ----
wb = load_workbook(FILE, data_only=True)
ws_in = wb[INPUT_SHEET]  # ← シート名が違うならここ上の変数を直す

# data_only=True の意味
# 数式セルがあるとき、式そのものではなく**直近保存時の“計算結果の値”**を返す設定です。
# 例：セルに =SUM(A1:A3) が入っている場合
# data_only=False → '=SUM(A1:A3)'（式文字列）
# data_only=True → 42（前回Excelで計算して保存された数値）
# 注意：openpyxlは計算はしません。Excel側で再計算して保存していないと、None になることがあります。

# ws_in = wb[INPUT_SHEET]
# 役割：ブック wb の中から、タブ名（シート名）で目的のシートを取り出します。戻り値は Worksheet。
# 例：INPUT_SHEET = "テストシート" → wb["テストシート"] と同じ。

# ---- 3) カテゴリ別に合計金額を集計（B列=カテゴリ, D列=金額）----
totals = {}  # ふつうのdictでOK（キーが無ければ0から足す）
for r in range(2, ws_in.max_row + 1):  # 2行目から最終行まで
    cat = ws_in[f"B{r}"].value
    amt = ws_in[f"D{r}"].value
    if not cat:  # カテゴリ空はスキップ（最小ルール）
        continue
    amt = float(amt or 0)  # None/空は0として扱う、文字でもfloat化を試みる簡易処理
    totals[cat] = totals.get(cat, 0.0) + amt

# dict.get(key, default) は、「key が存在すればその値、無ければ default」を返します。
# ここでは まだそのカテゴリが現れていない最初の1回目 に 0.0 を返すために使っています。
# 1件目（初めて「食費」を見た）
# totals に "食費" が未登録 → totals.get("食費", 0.0) は 0.0
# 0.0 + 1200.0 → 1200.0
# totals["食費"] = 1200.0
# 2件目（再び「食費」が出た）
# すでに "食費": 1200.0 がある → .get("食費", 0.0) は 1200.0
# 1200.0 + 680.0 → 1880.0
# totals["食費"] = 1880.0（更新）

# ---- 4) 出力シートを作成（存在していたら消して作り直す）----
if OUTPUT_SHEET in wb.sheetnames:
    del wb[OUTPUT_SHEET]
ws_out = wb.create_sheet(OUTPUT_SHEET)

# ---- 5) 見出しを書き、カテゴリ合計を出力（合計の大きい順で並べる）----
ws_out["A1"] = "カテゴリ"
ws_out["B1"] = "合計金額"

row = 2
for cat, total in totals.items():
    ws_out[f"A{row}"] = cat
    ws_out[f"B{row}"] = round(total, 2)
    row += 1

# totals.items()
# これは辞書のすべての「キー」と「値」のペアを取り出すメソッドです。

# ws_out[f"A{row}"] = cat
# ws_out[f"B{row}"] = round(total, 2)
# row += 1
# つまり、上の表をExcelのA列・B列に転記していくイメージです。
# round(数値、 桁数) は Python標準の「四捨五入（丸め）」関数

# ---- 6) 保存（同じファイルに上書き保存）----
wb.save(FILE)
print("✅ 完了:", FILE.resolve())
print(f"   入力シート: {INPUT_SHEET}  → 出力シート: {OUTPUT_SHEET}")
