# ===============================================
# 目的：
#   既存の Excel(.xlsx) を Python (openpyxl) で開き、
#   特定セルへ値を書き込み、別名で保存する一連の基本フローを体験する。
#   ※ このファイルは「理解のための解説コメント」を増やしています。
# ===============================================

# ---- インポートはファイル先頭にまとめる（Lint/E402回避） ----
from pathlib import Path  # ファイルやフォルダの「場所（パス）」を扱う標準クラス
from datetime import datetime  # 現在日時を取得するため
from openpyxl import load_workbook, Workbook  # Excelブックの読み込み／新規作成に使う


# ---- ログ出力の小さなヘルパー（見やすさのため任意） ----
def log(*args):
    """シンプルなログ出力。printに '▶' を付けて見やすくするだけ。"""
    print("▶", *args)


# ---- 基準パスの決め方（超重要） ----
# どこから実行しても壊れないように、
# 「このスクリプトファイル（__file__）が置かれているフォルダ」を基準にする。
# 「このPythonファイル（例：run_excel_io_annotated.py）が入っているフォルダを、作業の基準点にする」
#
HERE: Path = Path(__file__).resolve().parent

# 読み取り元（既存ファイル）と保存先（別名）のパスを決める。
# いまは同じフォルダ(HERE)に置く想定。
# SRC（変数）「読み取る（入力）側のExcelファイル」へのパス（既存ファイルを開くため）
# OUT（変数）「書き込んで保存する（出力）側のExcelファイル」へのパス（編集結果を別名で保存するため）
SRC: Path = HERE / "sample_created.xlsx"  # 既存ファイル（なければ後で自動作成）
OUT: Path = HERE / "sample_created_edited.xlsx"  # 別名保存先（上書き事故防止）


def ensure_src_exists() -> None:
    """
    読み取り元の Excel が存在しない場合に、最低限の内容で新規作成する。
    学習・検証を止めないための安全策。
    """
    if SRC.exists():
        log("既存の sample_created.xlsx を使用します:", SRC)
        return

    log("sample_created.xlsx が無いので新規作成します:", SRC)
    wb = Workbook()  # 空のブックを作る（デフォルトで1枚シート付き）
    ws = wb.active  # 先頭シートを取り出す
    ws.title = "テストシート"  # シート名は任意（ここでは日本語）
    ws["A1"] = "こんにちは！Pythonから書き込みました。"  # A1に文字列を入れておく
    wb.save(SRC)  # ディスクに保存
    log("初期ファイル作成OK:", SRC)


def main() -> None:
    # ---- 1) パス確認（デバッグの基本） ----
    log("スクリプトの場所(HERE):", HERE)
    log("読み取り元(SRC):", SRC)
    log("保存先(OUT):", OUT)

    # ---- 2) 読み取り元が無ければ作成 ----
    ensure_src_exists()

    # ---- 3) 既存ブックを開く ----
    # load_workbook で .xlsx を開いて操作可能にする。
    # 先頭シート ws = wb.active でも OK。名前が分かっていれば wb["シート名"] の方が明示的。
    wb = load_workbook(SRC)
    ws = wb.active
    log("対象シート名:", ws.title)

    # ---- 4) セルの読み取り（.value で中身） ----
    # ここでは確認のために A1, B1 を表示してみる。
    # log()は出力を見やすくするためのヘルパー関数。print()と区別するため。
    # 実際には ws["A1"].value のようにしてセルの中身を取得する。
    log("実行前 A1:", ws["A1"].value, "/ 実行前 B1:", ws["B1"].value)

    # ---- 5) セルへの書き込み（代入するだけ） ----
    # openpyxl は「セルへ値を代入」でOK。型は入れた値に応じて自動（str, int, datetimeなど）。
    ws["B1"] = 12345  # 数値
    ws["D1"] = "こんにちは openpyxl!"  # 文字列
    ws["F1"] = datetime.now()  # 現在日時（Excelではシリアル値として保存）

    # 参考：数式を書きたい場合は "=..." の文字列を入れる（例：平均）
    # ws["G1"] = "=AVERAGE(B2:B6)"
    # → Python保存後に Excel を開いて再計算→保存、再度 Python で data_only=True で読む、が定石。

    # ---- 6) 別名で保存（上書き事故を避ける） ----
    # Excel アプリで元ファイルを開いたままだとロックされるので要注意（⌘Qで完全終了してから実行）。
    wb.save(OUT)
    log("✅ 保存完了:", OUT)

    # ---- 7) 保存内容の確認（再読み込み） ----
    # もう一度開いて B1 が 12345 になっていることを確認。
    wb2 = load_workbook(OUT)
    ws2 = wb2.active
    log("再読込後 B1:", ws2["B1"].value)


# ---- スクリプトのエントリーポイント（Pythonの定石） ----
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        # 何かあっても「どこで落ちたか」を見失わないための最低限の例外処理。
        import traceback

        print("❌ 失敗:", e)
        traceback.print_exc()
