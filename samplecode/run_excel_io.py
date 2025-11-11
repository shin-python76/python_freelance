# run_excel_io.py
from __future__ import annotations
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook


def log(*args):  # 見やすいログ関数
    print("▶", *args)


HERE = Path(__file__).resolve().parent  # このスクリプトと同じフォルダ
SRC = HERE / "sample_created.xlsx"  # 読み取り元
OUT = HERE / "sample_created_edited.xlsx"  # 別名保存先


def ensure_src_exists():
    """もし sample_created.xlsx が無ければ作っておく（初期化）"""
    if SRC.exists():
        log("既存の sample_created.xlsx を使用します:", SRC)
        return
    log("sample_created.xlsx が無いので新規作成します:", SRC)
    wb = Workbook()
    ws = wb.active
    ws.title = "テストシート"
    ws["A1"] = "こんにちは！Pythonから書き込みました。"
    wb.save(SRC)
    log("初期ファイル作成OK:", SRC)


def main():
    log("スクリプト場所:", HERE)
    log("読み取り元:", SRC)
    log("保存先:", OUT)

    ensure_src_exists()

    # 既存ブックを開く
    wb = load_workbook(SRC)
    ws = wb.active
    log("シート名:", ws.title)
    log("実行前 A1:", ws["A1"].value, "/ 実行前 B1:", ws["B1"].value)

    # 書き込み（常に上書き）
    ws["B1"] = 12345
    ws["D1"] = "こんにちは openpyxl!"
    ws["F1"] = datetime.now()

    # 別名保存
    wb.save(OUT)
    log("✅ 保存完了:", OUT)

    # 再読み込みで確認
    wb2 = load_workbook(OUT)
    ws2 = wb2.active
    log("再読込後 B1:", ws2["B1"].value)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        # 失敗時は理由を出す
        import traceback

        print("❌ 失敗:", e)
        traceback.print_exc()
